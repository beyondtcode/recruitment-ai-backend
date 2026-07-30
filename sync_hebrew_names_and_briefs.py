"""
One-off sync: Hebrew candidate names + missing AI briefs from BeyondCode.xlsx → Monday Main Hub.

Loads the Main Hub board once into an email index, then renames / uploads missing briefs.

Usage:
    python sync_hebrew_names_and_briefs.py --dry-run
    python sync_hebrew_names_and_briefs.py
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Literal

from dotenv import load_dotenv
from openpyxl import load_workbook

from crm_integration.monday_client import fetch_all_board_items
from services.monday_service import (
    AI_SUMMARY_COLUMN_ID,
    CHANGE_MULTIPLE_COLUMN_VALUES_MUTATION,
    _email_from_column,
    change_item_name,
    get_main_hub_board_id,
    normalize_email,
    post_graphql,
    resolve_column_id_by_type,
)

_BACKEND_ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = Path(
    r"c:\Users\user\Downloads\BeyondCode\BeyondCode\BeyondCode.xlsx"
)
REPORT_PATH = _BACKEND_ROOT / "sync_hebrew_names_and_briefs_report.json"
WRITE_DELAY_SECONDS = 0.05

logger = logging.getLogger(__name__)

RowStatus = Literal[
    "renamed",
    "name_unchanged",
    "brief_uploaded",
    "brief_already_present",
    "brief_missing_in_excel",
    "not_found",
    "skipped_no_email",
    "skipped_no_name",
    "skipped_dry_run",
    "error",
]


@dataclass
class ExcelCandidate:
    email: str
    first_name: str
    last_name: str
    brief: str
    candidate_number: str | None = None

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


@dataclass
class MondayCandidate:
    item_id: str
    name: str
    brief: str


@dataclass
class RowResult:
    email: str
    excel_name: str
    status: RowStatus
    item_id: str | None = None
    monday_name: str | None = None
    renamed: bool = False
    brief_uploaded: bool = False
    brief_already_present: bool = False
    error: str | None = None


@dataclass
class SyncReport:
    dry_run: bool
    excel_path: str
    board_id: str
    total_excel_rows: int = 0
    unique_emails: int = 0
    monday_items_indexed: int = 0
    renamed: int = 0
    name_unchanged: int = 0
    brief_uploaded: int = 0
    brief_already_present: int = 0
    brief_missing_in_excel: int = 0
    not_found: int = 0
    skipped_no_email: int = 0
    skipped_no_name: int = 0
    skipped_dry_run: int = 0
    errors: int = 0
    rows: list[dict[str, Any]] = field(default_factory=list)

    def add(self, result: RowResult, *, excel_brief: str = "") -> None:
        self.rows.append(asdict(result))
        if result.status == "error":
            self.errors += 1
            return
        if result.status == "not_found":
            self.not_found += 1
            return
        if result.status == "skipped_no_email":
            self.skipped_no_email += 1
            return
        if result.status == "skipped_no_name":
            self.skipped_no_name += 1
            return
        if result.status == "skipped_dry_run":
            self.skipped_dry_run += 1

        if result.item_id:
            if result.renamed:
                self.renamed += 1
            else:
                self.name_unchanged += 1

            if result.brief_uploaded:
                self.brief_uploaded += 1
            elif result.brief_already_present:
                self.brief_already_present += 1
            elif not excel_brief:
                self.brief_missing_in_excel += 1


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_index(headers: list[Any], name: str) -> int:
    for i, header in enumerate(headers):
        if _cell_str(header) == name:
            return i
    raise KeyError(f"Missing Excel column: {name!r}")


def load_excel_candidates(excel_path: Path) -> tuple[list[ExcelCandidate], int, int]:
    """Return (deduped candidates, raw row count, skipped_no_email count)."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))

    idx_last = _header_index(headers, "שם משפחה")
    idx_first = _header_index(headers, "שם פרטי")
    idx_email = _header_index(headers, "מייל")
    idx_brief = _header_index(headers, "תקציר")
    idx_num = None
    try:
        idx_num = _header_index(headers, "מספר מועמד")
    except KeyError:
        pass

    by_email: dict[str, ExcelCandidate] = {}
    raw_rows = 0
    skipped_no_email = 0

    for row in rows_iter:
        raw_rows += 1
        if not row or not any(_cell_str(c) for c in row):
            continue

        email_raw = _cell_str(row[idx_email] if idx_email < len(row) else None)
        if not email_raw or "@" not in email_raw:
            skipped_no_email += 1
            continue

        email = normalize_email(email_raw)
        first = _cell_str(row[idx_first] if idx_first < len(row) else None)
        last = _cell_str(row[idx_last] if idx_last < len(row) else None)
        brief = _cell_str(row[idx_brief] if idx_brief < len(row) else None)
        number = None
        if idx_num is not None and idx_num < len(row):
            number = _cell_str(row[idx_num]) or None

        existing = by_email.get(email)
        if existing is None:
            by_email[email] = ExcelCandidate(
                email=email,
                first_name=first,
                last_name=last,
                brief=brief,
                candidate_number=number,
            )
            continue

        if first:
            existing.first_name = first
        if last:
            existing.last_name = last
        if brief:
            existing.brief = brief
        if number:
            existing.candidate_number = number

    wb.close()
    return list(by_email.values()), raw_rows, skipped_no_email


def _column_text(item: dict[str, Any], column_id: str) -> str:
    for col in item.get("column_values") or []:
        if col.get("id") == column_id:
            return _cell_str(col.get("text"))
    return ""


def _column_dict(item: dict[str, Any], column_id: str) -> dict[str, Any]:
    for col in item.get("column_values") or []:
        if col.get("id") == column_id:
            return col
    return {}


async def build_monday_email_index(board_id: str) -> dict[str, MondayCandidate]:
    email_column_id = await resolve_column_id_by_type(board_id, "email")
    logger.info(
        "Fetching all Main Hub items (email=%s, brief=%s)...",
        email_column_id,
        AI_SUMMARY_COLUMN_ID,
    )
    items = await fetch_all_board_items(
        board_id,
        [email_column_id, AI_SUMMARY_COLUMN_ID],
    )
    index: dict[str, MondayCandidate] = {}
    for item in items:
        email = _email_from_column(_column_dict(item, email_column_id))
        if not email:
            continue
        index[email] = MondayCandidate(
            item_id=str(item["id"]),
            name=_cell_str(item.get("name")),
            brief=_column_text(item, AI_SUMMARY_COLUMN_ID),
        )
    logger.info(
        "Indexed %d Monday items with email (from %d total items)",
        len(index),
        len(items),
    )
    return index


async def _write_ai_summary(board_id: str, item_id: str, brief: str) -> None:
    column_values = json.dumps({AI_SUMMARY_COLUMN_ID: brief})
    await post_graphql(
        CHANGE_MULTIPLE_COLUMN_VALUES_MUTATION,
        {
            "boardId": board_id,
            "itemId": item_id,
            "columnValues": column_values,
        },
        column_ids=[AI_SUMMARY_COLUMN_ID],
    )


async def process_candidate(
    candidate: ExcelCandidate,
    monday: MondayCandidate | None,
    *,
    board_id: str,
    dry_run: bool,
) -> RowResult:
    excel_name = candidate.full_name
    if not excel_name:
        return RowResult(
            email=candidate.email,
            excel_name="",
            status="skipped_no_name",
        )

    if monday is None:
        return RowResult(
            email=candidate.email,
            excel_name=excel_name,
            status="not_found",
        )

    try:
        monday_name = monday.name.strip()
        needs_rename = monday_name != excel_name
        has_excel_brief = bool(candidate.brief)
        monday_brief_empty = not monday.brief
        needs_brief = has_excel_brief and monday_brief_empty

        if dry_run:
            return RowResult(
                email=candidate.email,
                excel_name=excel_name,
                status="skipped_dry_run",
                item_id=monday.item_id,
                monday_name=monday_name,
                renamed=needs_rename,
                brief_uploaded=needs_brief,
                brief_already_present=bool(monday.brief),
            )

        renamed = False
        if needs_rename:
            await change_item_name(monday.item_id, excel_name, board_id=board_id)
            renamed = True
            await asyncio.sleep(WRITE_DELAY_SECONDS)

        brief_uploaded = False
        brief_already_present = bool(monday.brief)
        if needs_brief:
            await _write_ai_summary(board_id, monday.item_id, candidate.brief)
            brief_uploaded = True
            await asyncio.sleep(WRITE_DELAY_SECONDS)

        if renamed:
            primary_status: RowStatus = "renamed"
        elif brief_uploaded:
            primary_status = "brief_uploaded"
        elif brief_already_present:
            primary_status = "brief_already_present"
        elif not has_excel_brief:
            primary_status = "brief_missing_in_excel"
        else:
            primary_status = "name_unchanged"

        return RowResult(
            email=candidate.email,
            excel_name=excel_name,
            status=primary_status,
            item_id=monday.item_id,
            monday_name=monday_name,
            renamed=renamed,
            brief_uploaded=brief_uploaded,
            brief_already_present=brief_already_present,
        )
    except Exception as exc:
        logger.exception("Failed processing %s", candidate.email)
        return RowResult(
            email=candidate.email,
            excel_name=excel_name,
            status="error",
            item_id=monday.item_id if monday else None,
            error=str(exc),
        )


async def run_sync(
    *,
    excel_path: Path,
    dry_run: bool,
    limit: int | None = None,
) -> SyncReport:
    board_id = get_main_hub_board_id()
    candidates, raw_rows, skipped_no_email = load_excel_candidates(excel_path)
    if limit is not None:
        candidates = candidates[:limit]

    monday_index = await build_monday_email_index(board_id)

    report = SyncReport(
        dry_run=dry_run,
        excel_path=str(excel_path),
        board_id=board_id,
        total_excel_rows=raw_rows,
        unique_emails=len(candidates),
        monday_items_indexed=len(monday_index),
        skipped_no_email=skipped_no_email,
    )

    logger.info(
        "Processing %d unique emails against %d Monday emails (dry_run=%s)",
        len(candidates),
        len(monday_index),
        dry_run,
    )

    for i, candidate in enumerate(candidates, start=1):
        result = await process_candidate(
            candidate,
            monday_index.get(candidate.email),
            board_id=board_id,
            dry_run=dry_run,
        )
        report.add(result, excel_brief=candidate.brief)

        if i % 500 == 0 or i == len(candidates):
            logger.info(
                "Progress %d/%d renamed=%d briefs=%d not_found=%d errors=%d",
                i,
                len(candidates),
                report.renamed,
                report.brief_uploaded,
                report.not_found,
                report.errors,
            )

    return report


def _persist_report(report: SyncReport) -> None:
    REPORT_PATH.write_text(
        json.dumps(asdict(report), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _print_summary(report: SyncReport) -> None:
    print("\n=== Hebrew names + briefs sync ===")
    print(f"Excel:                 {report.excel_path}")
    print(f"Board:                 {report.board_id}")
    print(f"Dry run:               {report.dry_run}")
    print(f"Excel rows:            {report.total_excel_rows}")
    print(f"Unique emails:         {report.unique_emails}")
    print(f"Monday indexed:        {report.monday_items_indexed}")
    print(f"Skipped (no email):    {report.skipped_no_email}")
    print(f"Skipped (no name):     {report.skipped_no_name}")
    label = "Would rename" if report.dry_run else "Renamed"
    print(f"{label + ':':<23}{report.renamed}")
    print(f"Name unchanged:        {report.name_unchanged}")
    label_b = "Would upload brief" if report.dry_run else "Brief uploaded"
    print(f"{label_b + ':':<23}{report.brief_uploaded}")
    print(f"Brief already present: {report.brief_already_present}")
    print(f"Brief missing in excel:{report.brief_missing_in_excel}")
    print(f"Not found on Monday:   {report.not_found}")
    print(f"Errors:                {report.errors}")
    print(f"Report:                {REPORT_PATH}")


def main() -> int:
    load_dotenv(_BACKEND_ROOT / ".env")

    parser = argparse.ArgumentParser(
        description="Sync Hebrew candidate names and missing briefs from Excel to Monday.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview renames/brief uploads without writing to Monday.",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help=f"Path to BeyondCode.xlsx (default: {DEFAULT_EXCEL_PATH})",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Process only the first N unique emails (for testing).",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")

    if not args.excel.exists():
        print(f"Excel file not found: {args.excel}", file=sys.stderr)
        return 1

    report: SyncReport | None = None
    try:
        report = asyncio.run(
            run_sync(
                excel_path=args.excel,
                dry_run=args.dry_run,
                limit=args.limit,
            )
        )
    except Exception as exc:
        print(f"Sync failed: {exc}", file=sys.stderr)
        logger.exception("Sync failed")
        return 1
    finally:
        if report is not None:
            _persist_report(report)

    if report is None:
        return 1

    _print_summary(report)
    return 1 if report.errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
